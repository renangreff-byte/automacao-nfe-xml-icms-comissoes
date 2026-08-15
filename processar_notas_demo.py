"""Versão pública e anonimizada para demonstração de portfólio."""

from __future__ import annotations

import json
import re
import shutil
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


BASE = Path(__file__).resolve().parent
ENTRADA = BASE / "entrada"
SAIDA = BASE / "saida" / "relatorio_nfe_demo.xlsx"
NS = {"nfe": "http://www.portalfiscal.inf.br/nfe"}


def carregar_json(nome: str) -> dict:
    with (BASE / nome).open(encoding="utf-8") as arquivo:
        return json.load(arquivo)


def texto(elemento: ET.Element | None, caminho: str, padrao: str = "") -> str:
    if elemento is None:
        return padrao
    encontrado = elemento.find(caminho, NS)
    return (encontrado.text or "").strip() if encontrado is not None else padrao


def decimal(valor: str) -> Decimal:
    try:
        return Decimal(valor)
    except Exception:
        return Decimal("0")


def classificar_produto(descricao: str) -> str | None:
    normalizado = descricao.upper()
    if "ARLA" in normalizado:
        return "ARLA"
    if "DIESEL" in normalizado:
        return "DIESEL"
    return None


def extrair_placa(texto_livre: str, motoristas: dict) -> str:
    for placa in motoristas:
        if placa.upper() in texto_livre.upper():
            return placa
    correspondencia = re.search(r"PLACA\s*[:=-]?\s*([A-Z0-9-]{6,8})", texto_livre.upper())
    return correspondencia.group(1) if correspondencia else "NAO IDENTIFICADA"


def processar_xml(caminho: Path, filiais: dict, motoristas: dict) -> tuple[dict, list[dict]]:
    raiz = ET.parse(caminho).getroot()
    inf_nfe = raiz.find(".//nfe:infNFe", NS)
    if inf_nfe is None:
        raise ValueError("Estrutura infNFe não encontrada")

    numero = texto(inf_nfe, "nfe:ide/nfe:nNF")
    emissao = texto(inf_nfe, "nfe:ide/nfe:dhEmi") or texto(inf_nfe, "nfe:ide/nfe:dEmi")
    data = datetime.fromisoformat(emissao).date() if emissao else None
    cnpj_filial = texto(inf_nfe, "nfe:dest/nfe:CNPJ")
    filial = filiais.get(cnpj_filial, "FILIAL NAO MAPEADA")
    emitente = texto(inf_nfe, "nfe:emit/nfe:xNome")
    chave = inf_nfe.attrib.get("Id", "").removeprefix("NFe")
    icms = decimal(texto(inf_nfe, "nfe:total/nfe:ICMSTot/nfe:vICMS"))
    valor_nf = decimal(texto(inf_nfe, "nfe:total/nfe:ICMSTot/nfe:vNF"))
    complemento = texto(inf_nfe, "nfe:infAdic/nfe:infCpl")

    nota = {
        "filial": filial,
        "mes": data.strftime("%Y-%m") if data else "SEM DATA",
        "numero": numero,
        "chave": chave,
        "data": data,
        "emitente": emitente,
        "icms": icms,
        "valor_nf": valor_nf,
    }

    itens = []
    for det in inf_nfe.findall("nfe:det", NS):
        produto = det.find("nfe:prod", NS)
        descricao = texto(produto, "nfe:xProd")
        categoria = classificar_produto(descricao)
        if not categoria:
            continue
        texto_item = " ".join([descricao, texto(produto, "nfe:infAdProd"), complemento])
        placa = extrair_placa(texto_item, motoristas)
        cadastro = motoristas.get(placa, {})
        valor = decimal(texto(produto, "nfe:vProd"))
        taxa = decimal(str(cadastro.get("taxa_comissao", 0)))
        itens.append(
            {
                **nota,
                "produto": categoria,
                "litros": decimal(texto(produto, "nfe:qCom")),
                "valor": valor,
                "placa": placa,
                "motorista": cadastro.get("nome", "NAO IDENTIFICADO"),
                "tipo_frota": cadastro.get("tipo", "NAO CLASSIFICADO"),
                "taxa_comissao": taxa,
                "comissao": (valor * taxa).quantize(Decimal("0.01")),
            }
        )
    return nota, itens


def estilizar_cabecalho(ws) -> None:
    preenchimento = PatternFill("solid", fgColor="1F4E78")
    for celula in ws[1]:
        celula.font = Font(color="FFFFFF", bold=True)
        celula.fill = preenchimento
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for coluna in ws.columns:
        largura = min(max(len(str(c.value or "")) for c in coluna) + 2, 45)
        ws.column_dimensions[coluna[0].column_letter].width = largura


def gerar_excel(notas: list[dict], itens: list[dict]) -> None:
    workbook = Workbook()
    resumo = workbook.active
    resumo.title = "PREVIA_ICMS"
    resumo.append(["MES", "FILIAL", "QUANTIDADE_NF", "VALOR_NF", "PREVIA_ICMS"])

    consolidado = defaultdict(lambda: {"notas": 0, "valor": Decimal("0"), "icms": Decimal("0")})
    for nota in notas:
        chave = (nota["mes"], nota["filial"])
        consolidado[chave]["notas"] += 1
        consolidado[chave]["valor"] += nota["valor_nf"]
        consolidado[chave]["icms"] += nota["icms"]
    for (mes, filial), totais in sorted(consolidado.items()):
        resumo.append([mes, filial, totais["notas"], float(totais["valor"]), float(totais["icms"])])

    abastecimentos = workbook.create_sheet("ABASTECIMENTOS")
    abastecimentos.append(["DATA", "NF", "CHAVE", "FILIAL", "EMITENTE", "PRODUTO", "LITROS", "VALOR", "PLACA", "MOTORISTA", "TIPO"])
    for item in itens:
        abastecimentos.append([item["data"], item["numero"], item["chave"], item["filial"], item["emitente"], item["produto"], float(item["litros"]), float(item["valor"]), item["placa"], item["motorista"], item["tipo_frota"]])

    comissoes = workbook.create_sheet("COMISSOES")
    comissoes.append(["DATA", "NF", "MOTORISTA", "PLACA", "PRODUTO", "VALOR_BASE", "TAXA", "COMISSAO"])
    for item in itens:
        if item["tipo_frota"] == "TERCEIRO":
            comissoes.append([item["data"], item["numero"], item["motorista"], item["placa"], item["produto"], float(item["valor"]), float(item["taxa_comissao"]), float(item["comissao"])])

    for planilha in workbook.worksheets:
        estilizar_cabecalho(planilha)
    SAIDA.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(SAIDA)


def main() -> None:
    filiais = carregar_json("filiais_exemplo.json")
    motoristas = carregar_json("motoristas_exemplo.json")
    arquivos = sorted(ENTRADA.glob("*.xml"))
    if not arquivos:
        exemplo = BASE / "nfe_demonstracao.xml"
        ENTRADA.mkdir(exist_ok=True)
        shutil.copy2(exemplo, ENTRADA / exemplo.name)
        arquivos = sorted(ENTRADA.glob("*.xml"))
        print("Nenhum XML encontrado; usando automaticamente o exemplo fictício.")

    notas, itens, erros = [], [], []
    for arquivo in arquivos:
        try:
            nota, itens_nota = processar_xml(arquivo, filiais, motoristas)
            notas.append(nota)
            itens.extend(itens_nota)
        except Exception as erro:
            erros.append(f"{arquivo.name}: {erro}")

    gerar_excel(notas, itens)
    print(f"Relatório criado: {SAIDA}")
    print(f"XMLs processados: {len(notas)} | Itens: {len(itens)} | Erros: {len(erros)}")
    for erro in erros:
        print(f"- {erro}")


if __name__ == "__main__":
    main()
